"""
Excel export using openpyxl with styled headers and auto-width columns.
"""
from __future__ import annotations

import io
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.schemas import CardResult


_HEADERS = ["Image", "Name", "Company", "Emails", "Phones", "Address", "Confidence", "Method"]
_HEADER_COLOR = "1E293B"   # Dark slate


def generate_excel(results: List[CardResult]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Cards"

    # ── Header row ────────────────────────────────────────────────────
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=_HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # ── Data rows ────────────────────────────────────────────────────
    for row_idx, r in enumerate(results, start=2):
        row_data = [
            r.image,
            r.name,
            r.company,
            "; ".join(r.emails),
            "; ".join(r.phones),
            r.address,
            f"{r.confidence:.0%}",
            r.method.value,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F1F5F9")

    # ── Auto column widths ────────────────────────────────────────────
    col_widths = [22, 22, 28, 35, 25, 45, 12, 14]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header ────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
